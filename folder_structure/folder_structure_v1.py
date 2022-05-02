import os
import openpyxl
import re


class ExcelReader:

    def __init__(self, excel_file):
        self.project_structure = {}
        self.row_datas = []
        self.column_datas = []
        self.parent_columns = set()
        self.sheets_to_read = ['MAIN', 'LISTS', '<ASSET>', '<SHOT>']

        self.work_book = openpyxl.load_workbook(excel_file)
        self.read_sheet('MAIN')

        self.project_structure = {}

    def read_sheet(self, sheet_name):
        main_sheet = self.work_book[sheet_name]
        start_range, end_range = main_sheet.calculate_dimension().split(':')

        first_column, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        last_column, row_range = re.findall(r'(\w+?)(\d+)', end_range)[0]

        for cell in main_sheet[first_column]:
            if cell.value is not None:
                self.parent_columns.add(cell.value)

        for parent in self.parent_columns:
            for row in main_sheet.rows:
                row_list = []

                for cell in row:
                    # print(cell.row)
                    # print(cell.column)
                    # print(cell)
                    ls = [cell.value, cell.row, cell.column]
                    if cell.value is not None:
                        self.check_next_column_data(cell.column)
                        row_list.append(ls)
                if row_list:
                    print(row_list)
                    self.row_datas.append(row_list)

            print(self.row_datas)

            ls_data = self.check_next_column_data()
            self.project_structure[parent]

        # print(self.project_structure)

    def check_next_column_data(self, column):
        return []



    # def check_child_element(self, row_datas):
    #     row_length = len(row_datas)
    #     print(row_length)
    #     data_dict = {}
    #     for i in range(row_length):
    #         row = row_datas[i][0]
    #         print("row {}".format(row))
    #         key = row[0]
    #         current_row = row[1]
    #         current_column = row[2]
    #         while self.check_next_cell(current_row, current_column):
    #             next_cell = self.check_next_cell(current_row, current_column)
    #             if next_cell:
    #                 print("next_cell {}".format(next_cell))
    #                 data_dict[key] = next_cell[0]
    #
    #         print(key, current_row, current_column)
    #     return data_dict
    #
    # def check_next_cell(self, row, column):
    #     next_row = row + 1
    #     next_column = column + 1
    #     for data in self.row_datas:
    #         if data[0][1] == next_row and data[0][2] == column:
    #             print("line 72 {}".format(data[0][0]))
    #             return data
    #         elif data[0][1] == next_row and data[0][2] == next_column:
    #             print("line 73 {}".format(data[0][0]))
    #             return data
    #         else:
    #             return []
    #

    def check_empty_row(self, row):
        row_list = []
        for cell in row:
            # print(cell.row)
            # print(cell.column)
            # print(cell)
            row_list.append(cell.value)
        # Ignores the empty rows
        if set(row_list) == set([None]):
            return False
        else:
            return True

    #
    # def
    #     start_range, end_range = self.main_sheet.calculate_dimension().split(':')
    #
    #     first_column, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
    #     last_column, row_range = re.findall(r'(\w+?)(\d+)', end_range)[0]
    #
    #     for cell in self.main_sheet[first_column]:
    #         if cell.value is not None:
    #             self.parent_columns.add(cell.value)
    #
    #     for parent in self.parent_columns:
    #         self.project_structure[parent] = self.read_columns(first_column)
    #
    #     print(self.project_structure)

    def read_child_columns_data(self, child_column):
        column_datas = []
        for cell in self.main_sheet[child_column]:
            key = cell.value
            if key is not None:
                cell_data = [cell.value, cell.row, cell.column]
                column_datas.append(cell_data)

        return column_datas

    def get_next_column_data(self, row, column):
        cell_value = self.main_sheet.cell(row=row, column=column).value
        return cell_value

    def read_columns(self, parent_column):

        child_column = chr(ord(parent_column) + 1)
        child_column_data = {}
        child_datas = self.read_child_columns_data(child_column)
        for child in child_datas:
            # print("child data {}".format(child))

            next_row = child[1] + 1
            next_column = child[2] + 1
            datas = []
            while self.get_next_column_data(next_row, next_column) is not None:
                print("child_row - {} next_row {}".format(child[1], next_row))
                print("child_column - {} next_column {}".format(child[2], next_column))
                next_cell_value = self.get_next_column_data(next_row, next_column)

                if next_cell_value is None:
                    break
                datas.append(next_cell_value)
                next_row += 1

            print("datas {}".format(datas))

            # if self.get_next_column_data(next_row, next_column) is not None:
            #     print("child_column {}".format(child_column))
            #     self.read_columns(child_column, datas)

            child_column_data[child[0]] = datas
        return child_column_data


excel_reader = ExcelReader("template.xlsx")

'''

        if ord(child_column) <= ord(self.last_column):
            for cell in self.main_sheet[child_column]:
                key = cell.value
                if key is not None:
                    cell_data = [cell.value]
                    child_column_data.append(cell_data)
                    if key not in child_column_data_dict:
                        child_column_data_dict[cell.value] = cell_data
                    elif type(child_column_data_dict[cell.value]) == list:
                        child_column_data_dict[cell.value].append(cell_data)
                    else:
                        child_column_data_dict[cell.value] = [child_column_data_dict[cell.value], cell_data]
                    nxt_column = cell.column + 1
                    next_row = cell.row + 1
                    next_cell_data = []
                    for i in range(int(self.row_range)):
                        next_cell = self.main_sheet.cell(row=next_row + i, column=nxt_column).value
                        # print(" cell : {}{} = {}".format( chr(ord(child_column)+1), next_row + i, next_cell))

                        if next_cell is None:
                            self.read_child_columns(child_column)
                            break
                        next_cell_data.append(next_cell)

                    child_column_data.append(next_cell_data)

                    child_column_data_dict[cell.value] = child_column_data
                    print("child_column_data_dict[cell.value] {}".format(child_column_data_dict[cell.value]))
                    # print("next_cell_data {}".format(next_cell_data))




for column in self.main_sheet.iter_cols(values_only=True):
    if not set(column) == set([None]):
        column_datas.append(column)


for row in main_sheet.rows:
    row_list = []
    for cell in row:
        # print(cell.row)
        # print(cell.column)
        # print(cell)
        row_list.append(cell.value)
    # Ignores the empty rows
    if not set(row_list) == set([None]):
        row_datas.append(row_list)

print("column_data --- {}".format(column_datas))
print("row_datas --- {}".format(row_datas))
# folder_dict = {}
# for row_data in row_datas:
#     for data in row_data:
#         if data is not None:
#
#             print(data)

'''
