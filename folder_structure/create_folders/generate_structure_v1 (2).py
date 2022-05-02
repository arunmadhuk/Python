import openpyxl
from openpyxl.utils import column_index_from_string
import re
import json


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col):
        self.value = value
        self.col = col


# Example Excel reader class
class ExcelReader:
    def __init__(self, excel_file="template.xlsx"):
        # This should of course be read from the file
        self.work_book = openpyxl.load_workbook(excel_file)

        main_sheet = self.work_book['MAIN']

        start_range, end_range = main_sheet.calculate_dimension().split(':')

        first_col, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        last_col, row_end = re.findall(r'(\w+?)(\d+)', end_range)[0]
        max_column_idx = column_index_from_string(last_col)
        first_column_idx = column_index_from_string(first_col)

        self.values = []
        # DEREK: should start from cell [first_col, row_start], not from [0, 0]
        for row in main_sheet.iter_rows(min_row=int(row_start), max_row=int(row_end), max_col=max_column_idx):
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    col = cell.column - first_column_idx
                    self.values.append(Item(value, col))

        self.counter = 0

    def load_sheet(self, sheet_name):
        data_list = []
        selected_sheet = self.work_book[sheet_name]
        data_range = selected_sheet.calculate_dimension()
        start_range, end_range = data_range.split(':')

        first_col, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        last_col, row_end = re.findall(r'(\w+?)(\d+)', end_range)[0]
        min_col_idx = column_index_from_string(first_col)
        max_col_idx = column_index_from_string(last_col)

        # DEREK: What you did doesn't work if there are empty cells.
        # => Shoud do something like:
        #   - have a list of dictionaries, 1 dictionary for each row (like you did)
        #   - then read the first row, and for each one read the full column
        #   - then for each value in column, add the entry key/value in the corresponding dictionary
        # Pseudo code:
        #    (not 100% sure it is correct but it is the main idea)
        #    list = []
        #    for each cell in the first row that is not empty
        #        key = cell.value
        #        i = 0
        #        read the rest of the column
        #        for each cell in the column
        #            if len(list) <= i:
        #               append empty dictionary to list
        #            add key/value in the dictionary in list[i] (even if empty value)
        #            i += 1

        header_row_idx = start_range + ":" + last_col + row_start
        key_list = []
        for header_rows in selected_sheet[header_row_idx]:
            for header_row in header_rows:
                key_list.append(header_row.value)

        for row in selected_sheet.iter_rows(min_row=int(row_start) + 1, max_row=int(row_end),
                                            min_col=min_col_idx, max_col=max_col_idx):
            new_dict = {}
            for i in range(len(key_list)):
                cell = row[i]
                new_dict[key_list[i]] = cell.value

            data_list.append(new_dict)

        return data_list

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def reset(self):
        self.counter = 0


def update_dictionary(data_lists, key, value='name'):
    updated_dict = {}
    key = re.sub('\W+', '', key)
    for data_list in data_lists:
        # print(asset_data)
        data_key = data_list[key]
        data_value = data_list[value]
        value_dict = dict()
        value_dict[data_value] = {}
        if data_key in updated_dict.keys():
            updated_dict[data_key].update(value_dict)
        else:
            updated_dict[data_key] = value_dict

    return updated_dict


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            # DEREK: don't change "prev", we need to keep the current one
            # prev = next
            next = excel_reader.get_next()

    return None


# Usage example
d = {}

excel_reader = ExcelReader()
root = excel_reader.get_next()

generate_dictionary(d, root)
print(" Before Asset type update d:  -- {}".format(d))

# DEREK: should not have anything hardcoded, it is supposed to be fully dynamic (if not, there is no point in having a script...)
# Also, you check for example if there is the [short_type] element, and if not you skip the whole "asset" part. This doesn't work, [short_type] doesn't have to be there.
# => The process is here is much more complex, we need to handle several levels of tags with maybe different parents.
# I will do it.

short_key = "[short_type]"
dict_keys = set()

if short_key in d["{ROOT}"]['data']['assets'].keys():
    asset_data_list = excel_reader.load_sheet("<ASSET>")
    print("asset_data_list {}".format(asset_data_list))
    data_dict = update_dictionary(asset_data_list, short_key)
    d["{ROOT}"]['data']['assets'].update(data_dict)
    d["{ROOT}"]['data']['assets'].pop(short_key, None)

sequence_key = "[sequence]"
if sequence_key in d["{ROOT}"]['shots'].keys():
    shot_data_list = excel_reader.load_sheet("<SHOT>")
    data_dict = update_dictionary(shot_data_list, sequence_key)
    d["{ROOT}"]['shots'].update(data_dict)
    d["{ROOT}"]['shots'].pop(sequence_key, None)

print(" After Asset & Shot type update d:  -- {}".format(d))

with open('output.json', 'w') as json_file:
    json.dump(d, json_file, indent=4)
