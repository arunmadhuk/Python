import openpyxl
import json


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col, row):
        self.value = value
        self.col = col
        self.row = row


# Excel reader class
class ExcelReader:
    # TODO: add main sheet name as param?
    def __init__(self, excel_file):
        self.counter = 0
        self.row_counter = 0
        self.values = []
        self.types_sheet_data = []
        self.types_sheet_header = []
        self.work_book = openpyxl.load_workbook(excel_file,
                                                read_only=True,
                                                keep_links=False)
        self.get_values()

    def get_values(self, sheet_name="MAIN"):
        if sheet_name not in self.work_book.sheetnames:
            return []

        sheet = self.work_book[sheet_name]
        data_range = sheet.calculate_dimension()
        data = sheet[data_range]

        # Check that first cell is not empty
        # TODO: should check if is "root" or "project" or something specific...
        if not data[0][0].value:
            # TODO: print better error
            print("Error: first cell not root")
            return

        first_column_idx = data[0][0].column

        self.values = []
        for row in data:
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    col = cell.column - first_column_idx
                    row = cell.row
                    self.values.append(Item(value, col, row))
                    # Only 1 value per row, rest is ignored
                    break

        return self.values

    # TODO: keep read data in class? (avoid reading at every call)
    def load_sheet(self, sheet_name):
        if sheet_name not in self.work_book.sheetnames:
            return []

        selected_sheet = self.work_book[sheet_name]
        data_range = selected_sheet.calculate_dimension()
        data = selected_sheet[data_range]

        """
        # Get column titles
        key_list = []
        for column_title in data[0]:
            # Stop at first "None" value
            if not column_title.value:
                break
            key_list.append(column_title.value)

        # Create dictionary for each row of table
        data_list = []
        for row in data[1:]:
            # Stop at first row starting with "None" value
            if not row[0].value:
                break

            new_dict = {}
            for i in range(len(key_list)):
                new_dict[key_list[i]] = row[i].value

            data_list.append(new_dict)
        """
        # Get column titles
        key_list = []
        first = True
        for column_title in data[0]:
            # Stop at first "None" value
            if not column_title.value:
                break

            # Special case for first column (ID)
            if not first:
                key_list.append(column_title.column_letter)
            else:
                key_list.append("[ID]")
                first = False

        nb_keys = len(key_list)

        # Create dictionary for each row of table
        data_list = []
        for row in data:
            # Stop at first row starting with "None" value
            if not row[0].value:
                break

            new_dict = {}
            for i in range(nb_keys):
                new_dict[key_list[i]] = row[i].value

            data_list.append(new_dict)

        return data_list

    def load_types_sheet(self, sheet_name="TYPES"):
        selected_sheet = self.work_book[sheet_name]
        data_range = selected_sheet.calculate_dimension()
        data = selected_sheet[data_range]
        # Get column titles
        for column_title in data[0]:
            # Stop at first "None" value
            if not column_title.value:
                break
            col = column_title.column
            row = column_title.row
            value = column_title.value
            self.types_sheet_header.append(Item(value, col, row))

        first_row_idx = data[0][0].row

        for row in data[1:]:
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    row = cell.row
                    col = cell.column
                    self.types_sheet_data.append(Item(value, col, row))

        type_dict = {}
        first_element = self.get_next_row_item()
        self.generate_types_dict(type_dict, first_element)

        for key in type_dict:
            nest_dict = type_dict[key]
            for nest_key in nest_dict.keys():
                next_item = self.get_next_row(nest_key, first_row_idx)
                nest_dict[nest_key] = next_item

        return type_dict

    def generate_types_dict(self, dic, prev):
        next = self.get_next_row_item()

        while next:

            if next.col < prev.col:
                return next

            if next.col == prev.col:
                dic[next.value] = {}
                prev = next
                next = self.get_next_row_item()

            elif next.col == prev.col + 1:
                dic[prev.value] = {next.value: {}}
                next = self.generate_types_dict(dic[prev.value], next)
            else:
                next = self.get_next_row_item()

        return None

    def get_next_row_item(self):
        item = None
        if self.row_counter < len(self.types_sheet_data):
            item = self.types_sheet_data[self.row_counter]
            self.row_counter += 1
        return item

    def check_next_row_element(self, row, col):
        row_items = []
        for val in self.types_sheet_data:
            if val.row == row and val.col >= col:
                row_items.append(val)

        return row_items

    def check_row_header(self, col, first_row_idx):
        for val in self.types_sheet_header:
            if val.row == first_row_idx and val.col == col:
                return val.value

    def get_next_row(self, val, first_row_idx):
        value_len = len(self.types_sheet_data)
        for i in range(value_len):
            if val == self.types_sheet_data[i].value and i + 1 <= value_len:
                row_item = {}
                next_row_item = self.types_sheet_data[i + 1]
                col = next_row_item.col
                row = next_row_item.row
                next_items = self.check_next_row_element(row, col)

                for next_item in next_items:
                    header = self.check_row_header(next_item.col, first_row_idx)
                    row_item[header] = next_item.value

        return row_item

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def reset(self):
        self.counter = 0
        self.row_counter = 0


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            next = excel_reader.get_next()

    return None


d = {}

excel_reader = ExcelReader("template.xlsx")
root = excel_reader.get_next()

generate_dictionary(d, root)
print(" d: {}".format(d))

"""
def get_tag_values(type, context):
    print("+++ get_tag_values +++")
    data = excel_reader.load_sheet(type)
    print("sheet data: {}".format(data))
    children = set()

    print("context: {}".format(context))


    # Get "link" entry in data (if exists)
    link_entry = None
    for index in range(len(data)):
        print("data[index]: {}".format(data[index]))
        if data[index]["[ID]"] == "[Link]":
            link_entry = data.pop(index)
            break
    print("link_entry: {}".format(link_entry))


    for entry in data:
        id = entry["[ID]"]
        if id[0] == "[" and id[-1] == "]":
            # Special cases => Future process?
            continue

        for pair in context:
            if pair[0][0] == "<" and pair[0][-1] == ">":
                dtype = None

                if link_entry:
                    for key in link_entry:
                        if link_entry[key] == pair[0]:
                            dtype = key
                            break
                    #else:
                    #    error?

                if dtype:
                    print("dtype: {}".format(dtype))
                    if pair[1] != entry[dtype]:
                        break


            elif pair[0] in entry.keys():
                print("entry[pair[0]]: " + entry[pair[0]])
                if pair[1] != entry[pair[0]]:
                    break
        else:
            children.add(id)

    return list(children)
"""


def get_tag_values(type, context):
    print("+++ get_tag_values +++")
    print("type: {}".format(type))
    data = excel_reader.load_sheet(type)
    print("sheet data: {}".format(data))
    children = set()

    print("context: {}".format(context))

    # Get "link" entry in data (if exists)
    link_entry = None
    for index in range(len(data)):
        print("data[index]: {}".format(data[index]))
        if data[index]["[ID]"] == "[Link]":
            link_entry = data.pop(index)
            break
    print("link_entry: {}".format(link_entry))

    for entry in data:
        id = entry["[ID]"]
        if id[0] == "[" and id[-1] == "]":
            # Special cases => Future process?
            continue

        for pair in context:

            link_type = pair[0]
            start = link_type.find("<")
            end = link_type.rfind(">")
            if start == -1 or start >= end:
                continue

            tt = link_type[start:end + 1]
            print("tt: {}".format(tt))

            # Context value in same table
            if tt == type:
                split = link_type[end + 1:].split(":")
                # TODO: check if > 2
                if len(split) > 1:
                    column = split[1]
                else:
                    # "ID" column
                    column = "[ID]"

                print("pair[1]: {}".format(pair[1]))
                print("entry[column]: {}".format(entry[column]))
                if pair[1] != entry[column]:
                    print("DIFFERENT VALUE - SKIP: {} != {}".format(pair[1], entry[column]))
                    break

            else:
                dtype = None

                print("pair[0]: {}".format(pair[0]))
                if link_entry:
                    for key in link_entry:
                        print("key: {}".format(key))
                        print("link_entry[key]: {}".format(link_entry[key]))
                        if link_entry[key] == pair[0]:
                            dtype = key
                            break
                    # else:
                    #    error?

                if dtype:
                    print("dtype: {}".format(dtype))
                    if pair[1] != entry[dtype]:
                        print("DIFFERENT VALUE - SKIP: {} != {}".format(pair[1], entry[dtype]))
                        break




        else:
            children.add(id)

    print("children: {}".format(children))
    return list(children)


"""
def get_values(dic, type, context):
    print("+++ get_values +++")
    values = set()
    print("TYPE: {}".format(type))
    for key in dic:
        if key[0] != "<" or key[-1] != ">":
            continue

        #TODO: unused?
        ## Get values
        #tag_values = get_tag_values(key, context)
        #print("tag_values - sub: {}".format(tag_values))

        data = excel_reader.load_sheet(key)
        sub_values = set()

        for entry in data:
            id = entry["[ID]"]
            if id[0] == "[" and id[-1] == "]":
                # Special cases => Future process?
                continue

            if type in entry:
                sub_values.add(entry[type])

        if not sub_values:
            #TODO: need to extend context?
            sub_values = get_values(dic[key], type, context)

        values.update(sub_values)

    return list(values)
"""
"""
def get_values(dic, key):
    print("+++ get_values +++")
    print("KEY: {}".format(key))

    values = set()

    split = key.split(":")
    if len(split) == 1:
        #TODO: log error/warning?
        return []

    #TODO: check that sheet name starts/ends with "<|>" (?)
    sheet_name = ":".join(split[:-1])
    type = split[-1]

    data = excel_reader.load_sheet(sheet_name)
    print("data: {}".format(data))
    for entry in data:
        id = entry["[ID]"]
        if id[0] == "[" and id[-1] == "]":
            # Special cases => Future process?
            continue

        if type in entry:
            values.add(entry[type])

    print("values: {}".format(values))
    return list(values)
"""


# def get_values(dic, key):
def get_values(dic, key, context):
    print("+++ get_values +++")
    print("KEY: {}".format(key))
    print("context: {}".format(context))

    values = set()

    split = key.split(":")
    if len(split) == 1:
        # TODO: check that sheet name starts/ends with "<|>" (?)
        sheet_name = key
        type = "[ID]"
    else:
        # TODO: check that sheet name starts/ends with "<|>" (?)
        sheet_name = ":".join(split[:-1])
        type = split[-1]

    data = excel_reader.load_sheet(sheet_name)
    print("data: {}".format(data))

    # Get "link" entry in data (if exists)
    link_entry = None
    for index in range(len(data)):
        print("data[index]: {}".format(data[index]))
        if data[index]["[ID]"] == "[Link]":
            link_entry = data.pop(index)
            break
    print("link_entry: {}".format(link_entry))

    for entry in data:
        id = entry["[ID]"]
        if id[0] == "[" and id[-1] == "]":
            # Special cases => Future process?
            continue

        if type in entry:
            print("type: {}".format(type))

            for pair in context:

                link_type = pair[0]
                start = link_type.find("<")
                end = link_type.rfind(">")
                if start == -1 or start >= end:
                    continue

                tt = link_type[start:end + 1]
                print("tt: {}".format(tt))

                # TODO: case "tt == type" cannot happen (?)
                ######## BEGIN
                """
                # Context value in same table
                if tt == type:
                    split = link_type[end+1:].split(":")
                    #TODO: check if > 2
                    if len(split) > 1:
                        column = split[1]
                    else:
                        # "ID" column
                        column = "[ID]"

                    print("pair[1]: {}".format(pair[1]))
                    print("entry[column]: {}".format(entry[column]))
                    if pair[1] != entry[column]:
                        print("DIFFERENT VALUE - SKIP: {} != {}".format(pair[1], entry[column]))
                        break

                else:
                """
                ######## MID
                if True:
                    ######## END
                    dtype = None

                    print("pair[0]: {}".format(pair[0]))
                    if link_entry:
                        for key in link_entry:
                            print("key: {}".format(key))
                            print("link_entry[key]: {}".format(link_entry[key]))
                            if link_entry[key] == pair[0]:
                                dtype = key
                                break
                        # else:
                        #    error?

                    if dtype:
                        print("dtype: {}".format(dtype))
                        if pair[1] != entry[dtype]:
                            print("DIFFERENT VALUE - SKIP: {} != {}".format(pair[1], entry[dtype]))
                            break
            else:
                values.add(entry[type])

    print("values: {}".format(values))
    return list(values)


# TODO: rename!
def fct(dic, context=[]):
    out_dic = {}

    for item in dic:
        print("item: {}".format(item))

        values = []

        # TODO: handle error cases
        if item[0] == "[" and item[-1] == "]":

            # fields = item[1:-1].split(":")
            fields = item[1:-1].split("=")
            if len(fields) == 2:
                values = [fields[1]]
                type = fields[0]
            else:
                type = fields[0]
                values = get_values(dic[item], type, context)
                # values = get_values(dic[item], type)

            print("VALUES: {}".format(values))

        elif item[0] == "<" and item[-1] == ">":

            type = item
            values = get_tag_values(type, context)
            print("tag_values: {}".format(values))

        else:
            values = [item]
            # type = item # or None?
            type = None

        # print("values, type: {}, {}".format(values, type))

        for value in values:
            out_dic[value] = {"type": type}

            # sub_dict = fct(dic[item])
            if type:
                sub_dict = fct(dic[item], context + [(type, value)])
            else:
                sub_dict = fct(dic[item], context)
            # print("sub_dict: {}".format(sub_dict))

            out_dic[value]["children"] = sub_dict

            # Remove type from items with children, or if lists ("[|]")
            if sub_dict or (item[0] == "[" and item[-1] == "]"):
                out_dic[value]["type"] = None

    # print("out_dic: {}".format(out_dic))
    return out_dic


dicc = fct(d)

with open('output.json', 'w') as json_file:
    json.dump(dicc, json_file, indent=4)

TYPES = excel_reader.load_types_sheet()
print("types_data: {}".format(TYPES))

'''
def output(dic):
    out_dic = {}
    for key in dic:
        type = dic[key]["type"]
        print("type: {}".format(type))

        """
        if type is in the sheet "TYPES" (column "Tag)
            get the corsponding "Ftrack object type"
            add key/vale "key[ftrack_object_type]: {}" to out_dic
            return out_dic
        """

        out_dic[key] = output(dic[key]["children"])


    return out_dic
'''

# TYPES = {
#     "ASSET": {
#         "#0": {
#         },
#         "#1": {
#             "D": "char"
#         },
#         "#2": {
#             "D": "prop"
#         },
#         "#3": {
#             "D": "env"
#         },
#         "#4": {
#             "D": "assembly"
#         },
#         "#5": {
#             "D": "scene"
#         }
#     },
#     "SHOT": {
#         "#6": {
#             "C": "full_cg"
#         },
#         "#7": {
#             "C": "fx"
#         },
#         "#8": {
#             "C": "anim"
#         }
#     },
#     "DOCUMENT": {
#         "#9": {
#             "A": "test1",
#             "C": "test2"
#         }
#     }
# }

FTRACK = excel_reader.load_sheet("FTRACK")
print("FTRACK: {}".format(FTRACK))


# TODO: use tags with/without "<|>"?
def get_types(dic):
    # print("dic: {}".format(dic))

    out_dic = {}

    for key in dic:
        type = dic[key]["type"]
        if type:
            type = type[1:-1]

            # TODO: get TYPES from Excel file
            if type in TYPES:
                print("type: {}".format(type))

                value = None
                tag = TYPES[type]
                for object_type in tag:
                    if not len(tag[object_type]):
                        #  No "constraints" => Default
                        if not value:
                            value = object_type
                    else:
                        data = excel_reader.load_sheet("<" + type + ">")
                        # print("data: {}".format(data))

                        for entry in data:
                            if entry["[ID]"] == key:
                                break
                        else:
                            # TODO: error?
                            print("Entry not found '{}'".format(key))
                            continue

                        fields = tag[object_type]
                        for field in fields:
                            """
                            print("field: {}".format(field))
                            print("value: {}".format(fields[field]))
                            print("entry: {}".format(entry[field]))
                            """

                            if fields[field] != entry[field]:
                                break
                        else:
                            print("Found match: '{}'".format(object_type))
                            value = object_type

                if not value:
                    raise Exception("No matching entry for type! Add default case if required")

                for object_type in FTRACK:
                    if object_type["[ID]"] == value:
                        break
                else:
                    # TODO: error?
                    print("Object type not found in FTRACK: '{}'".format(value))

                # TODO: handle differently
                # => check with OpenPype + Ftrack
                ########
                """
                for field in object_type:
                    if field[0] == "[" and field[-1] == "]":
                        continue
                    if not object_type[field]:
                        continue

                    value = value + "|" + object_type[field]
                """
                ########

                out_dic[key] = value
                continue

        out_dic[key] = get_types(dic[key]["children"])

    return out_dic


# final = output(dicc)
final = get_types(dicc)
print(final)
with open('output2.json', 'w') as json_file:
    json.dump(final, json_file, indent=4)

PYPE_TYPE = "[Pype]"
NON_PYPE_TYPE = "[Folder]"


def convert_types(dic):
    out_dic = {}
    for key in dic:
        print("dic[key] {}".format(dic[key]))
        if isinstance(dic[key], dict):
            ispype = False
            tmp_dic = convert_types(dic[key])
            print("tmp_dic {}".format(tmp_dic))
            for tmp in tmp_dic:
                if not tmp.endswith(NON_PYPE_TYPE):
                    ispype = True
            if ispype:
                # Replace: if any of the keys contain/endswith PYPE_TYPE
                out_dic[key + PYPE_TYPE] = tmp_dic
            else:
                out_dic[key + NON_PYPE_TYPE] = tmp_dic
        else:
            new_key = key + "[" + dic[key] + "]"
            out_dic[new_key] = 0
    return out_dic


new_dic = convert_types(final)
print(new_dic)

with open('output3.json', 'w') as json_file:
    json.dump(new_dic, json_file, indent=4)