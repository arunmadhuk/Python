import openpyxl
from openpyxl.utils import column_index_from_string
import re
import json


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col):
        self.value = value
        self.col = col


# Example Excel reader class
class ExcelReader:
    def __init__(self, excel_file="template.xlsx"):
        # This should of course be read from the file
        self.work_book = openpyxl.load_workbook(excel_file)

        main_sheet = self.work_book['MAIN']

        start_range, end_range = main_sheet.calculate_dimension().split(':')

        first_col, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        # last_col, row_range = re.findall(r'(\w+?)(\d+)', end_range)[0]

        first_column_idx = column_index_from_string(first_col)
        print("first column idx -- {}".format(first_column_idx))
        self.values = []
        for row in main_sheet.rows:
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    col = cell.column - first_column_idx
                    self.values.append(Item(value, col))

        self.counter = 0

    def load_shot_sheet(self):
        shot_list = []
        shot_sheet = self.work_book['<SHOT>']
        start_range, end_range = shot_sheet.calculate_dimension().split(':')

        first_col, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        last_col, row_end = re.findall(r'(\w+?)(\d+)', end_range)[0]
        max_column_idx = column_index_from_string(last_col)
        key_list = []

        for cell in shot_sheet[row_start]:
            if cell.value is not None:
                # print(cell.value)
                key_list.append(cell.value)
        # print(key_list)

        for row in shot_sheet.iter_rows(min_row=int(row_start) + 1, max_row=int(row_end), max_col=max_column_idx):
            new_dict = {}
            i = 0
            for cell in row:
                if cell.value is not None:
                    new_dict[key_list[i]] = cell.value
                    i = i + 1

            # print(new_dict)
            shot_list.append(new_dict)

        return shot_list

    def load_asset_sheet(self):
        asset_list = []
        asset_sheet = self.work_book['<ASSET>']

        start_range, end_range = asset_sheet.calculate_dimension().split(':')

        first_col, row_start = re.findall(r'(\w+?)(\d+)', start_range)[0]
        last_col, row_end = re.findall(r'(\w+?)(\d+)', end_range)[0]
        max_column_idx = column_index_from_string(last_col)
        key_list = []
        # print(asset_sheet[row_start])
        for cell in asset_sheet[row_start]:
            if cell.value is not None:
                # print(cell.value)
                key_list.append(cell.value)
        # print(key_list)
        for row in asset_sheet.iter_rows(min_row=int(row_start) + 1, max_row=int(row_end), max_col=max_column_idx):
            new_dict = {}
            i = 0
            for cell in row:
                if cell.value is not None:
                    new_dict[key_list[i]] = cell.value
                    # print("Asset Sheet ---> cell value {}".format(cell.value))
                    i = i + 1

            # print(new_dict)
            asset_list.append(new_dict)

        return asset_list

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def reset(self):
        self.counter = 0


def update_dictionary(data_lists, key, value='name'):
    updated_dict = {}
    key = re.sub('\W+','',key)
    for data_list in data_lists:
        # print(asset_data)
        data_key = data_list[key]
        data_value = data_list[value]
        value_dict = dict()
        value_dict[data_value] = {}
        if data_key in updated_dict.keys():
            updated_dict[data_key].update(value_dict)
        else:
            updated_dict[data_key] = value_dict

    return updated_dict


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            prev = next
            next = excel_reader.get_next()

    return None


# Usage example
d = {}

excel_reader = ExcelReader()
root = excel_reader.get_next()

generate_dictionary(d, root)
print(" Before Asset type update d:  -- {}".format(d))

short_key = "[short_type]"
if short_key in d["{ROOT}"]['data']['assets'].keys():
    asset_data_list = excel_reader.load_asset_sheet()
    print("asset_data_list:  -- {}".format(asset_data_list))
    data_dict = update_dictionary(asset_data_list, short_key)
    d["{ROOT}"]['data']['assets'].update(data_dict)
    d["{ROOT}"]['data']['assets'].pop(short_key, None)

sequence_key = "[sequence]"
if sequence_key in d["{ROOT}"]['shots'].keys():
    shot_data_list = excel_reader.load_shot_sheet()
    print("shot_data_list:  -- {}".format(shot_data_list))
    data_dict = update_dictionary(shot_data_list, sequence_key)
    d["{ROOT}"]['shots'].update(data_dict)
    d["{ROOT}"]['shots'].pop(sequence_key, None)

# print(d["{ROOT}"].keys())
print(" After Asset & Shot type update d:  -- {}".format(d))

with open('output.json', 'w') as json_file:
    json.dump(d, json_file, indent=4)
