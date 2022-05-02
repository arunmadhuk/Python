import openpyxl
import json
import re


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col):
        self.value = value
        self.col = col


class RowItem:
    def __init__(self, value, row, col):
        self.value = value
        self.row = row
        self.col = col


# Excel reader class
class ExcelReader:
    # TODO: don't set excel file name a default
    # TODO: add main sheet name as param?
    def __init__(self, excel_file="template.xlsx"):
        self.counter = 0
        self.row_counter =0
        self.values = []
        self.header=[]
        self.row_values = []

        self.work_book = openpyxl.load_workbook(excel_file)
        self.get_values()

    def get_values(self, sheet_name="TYPES"):
        if sheet_name not in self.work_book.sheetnames:
            return []

        sheet = self.work_book[sheet_name]
        data_range = sheet.calculate_dimension()
        data = sheet[data_range]

        # Check that first cell is not empty
        # TODO: should check if is "root" or "project" or something specific...
        if not data[0][0].value:
            # TODO: print better error
            print("Error: first cell not root")
            # return

        first_column_idx = data[0][0].column
        self.first_row_idx = data[0][0].row
        print(first_column_idx)
        print(self.first_row_idx)

        for row_headers in sheet.iter_rows(min_row=self.first_row_idx, max_row=self.first_row_idx, min_col=first_column_idx,
                                           values_only=True):
            for row_header in row_headers:
                if row_header is None:
                    break
                else:
                    self.header.append(row_header)

        self.values = []
        for row in sheet.iter_rows(min_row=self.first_row_idx, min_col=first_column_idx,
                                   max_col=first_column_idx + 1):
            for cell in row:
                if cell.value is not None and cell.value not in self.header:
                    value = cell.value
                    col = cell.column - first_column_idx
                    self.values.append(Item(value, col))
                    # print(" {} {}".format(value, col))
                    # Only 1 value per row, rest is ignored
                    # break

        for row in data:
            for cell in row:
                if cell.value is not None :
                    value = cell.value
                    row = cell.row
                    col = cell.column
                    self.row_values.append(RowItem(value, row, col))
                    # Only 1 value per row, rest is ignored

        # print("lenght of row values {}".format(len(self.row_values)))
        #
        # for value in self.row_values:
        #     print(value.value, value.row, value.col)

        return self.values

    def load_sheet(self, sheet_name):
        if sheet_name not in self.work_book.sheetnames:
            return []

        selected_sheet = self.work_book[sheet_name]
        data_range = selected_sheet.calculate_dimension()
        data = selected_sheet[data_range]

        # Get column titles
        key_list = []
        first = True
        for column_title in data[0]:
            # Stop at first "None" value
            if not column_title.value:
                break

            # Special case for first column (ID)
            if not first:
                key_list.append(column_title.column_letter)
            else:
                key_list.append("[ID]")
                first = False

        nb_keys = len(key_list)

        # Create dictionary for each row of table
        data_list = []
        for row in data:
            # Stop at first row starting with "None" value
            if not row[0].value:
                break

            new_dict = {}
            for i in range(nb_keys):
                new_dict[key_list[i]] = row[i].value

            data_list.append(new_dict)

        return data_list

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def check_next_row_element(self, row, col):
        row_items = []
        for val in self.row_values:
            if val.row == row and val.col >= col:
                row_items.append(val)
        return row_items

    def check_row_header(self, col):
        for val in self.row_values:
            if val.row == self.first_row_idx and val.col == col:
                return val.value

    def get_next_row(self, val):
        for i in range(len(self.row_values)):

            if val == self.row_values[i].value:

                row_item = self.row_values[i+1]
                col = row_item.col
                row = row_item.row
                # print(col, row)
                # print(row_item.value)
                next_items = self.check_next_row_element(row, col)
                row_item = {}
                for next_item in next_items:
                    header = self.check_row_header(next_item.col)
                    row_item[header] = next_item.value

        return row_item


    def reset(self):
        self.counter = 0
        self.row_counter =0


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            next = excel_reader.get_next()

    return None


def update_dictionary(dic):

    for key in dic:
        nest_dict = dic[key]
        for nest_key in nest_dict.keys():
            next_item = excel_reader.get_next_row(nest_key)
            nest_dict[nest_key] = next_item

        # print(dict.keys())
    return dic

d = {}

excel_reader = ExcelReader()
root = excel_reader.get_next()

generate_dictionary(d, root)


dic = update_dictionary(d)

print(" dic: {}".format(dic))

with open('types_output.json', 'w') as json_file:
    json.dump(dic, json_file, indent=4)