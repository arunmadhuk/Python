import openpyxl
import json
import re


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col, row):
        self.value = value
        self.col = col
        self.row = row


# Excel reader class
class ExcelReader:
    # TODO: don't set excel file name a default
    # TODO: add main sheet name as param?
    def __init__(self, excel_file="template.xlsx"):
        self.counter = 0
        self.row_counter = 0
        self.values = []
        self.header = []
        self.row_values = []

        self.work_book = openpyxl.load_workbook(excel_file)
        self.get_values()

    def get_values(self, sheet_name="TYPES"):
        if sheet_name not in self.work_book.sheetnames:
            return []

        sheet = self.work_book[sheet_name]
        data_range = sheet.calculate_dimension()
        data = sheet[data_range]

        # Check that first cell is not empty
        # TODO: should check if is "root" or "project" or something specific...
        if not data[0][0].value:
            # TODO: print better error
            print("Error: first cell not root")
            # return

        first_column_idx = data[0][0].column
        self.first_row_idx = data[0][0].row
        # print(first_column_idx)
        # print(self.first_row_idx)

        for row_headers in sheet.iter_rows(min_row=self.first_row_idx, max_row=self.first_row_idx,
                                           min_col=first_column_idx):
            for row_header in row_headers:
                if row_header is None:
                    break
                else:
                    col = row_header.column
                    row = row_header.row
                    value = row_header.value
                    self.row_values.append(Item(value, col, row))
                    self.header.append(value)

        self.values = []
        for row in sheet.iter_rows(min_row=self.first_row_idx, min_col=first_column_idx):
            for cell in row:
                if cell.value is not None and cell.value not in self.header:
                    value = cell.value
                    row = cell.row
                    col = cell.column
                    self.values.append(Item(value, col, row))
                    # print(" {} {}".format(value, col))
                    # Only 1 value per row, rest is ignored
                    # break

        return self.values

    def load_sheet(self, sheet_name):
        if sheet_name not in self.work_book.sheetnames:
            return []

        selected_sheet = self.work_book[sheet_name]
        data_range = selected_sheet.calculate_dimension()
        data = selected_sheet[data_range]

        # Get column titles
        key_list = []
        first = True
        for column_title in data[0]:
            # Stop at first "None" value
            if not column_title.value:
                break

            # Special case for first column (ID)
            if not first:
                key_list.append(column_title.column_letter)
            else:
                key_list.append("[ID]")
                first = False

        nb_keys = len(key_list)

        # Create dictionary for each row of table
        data_list = []
        for row in data:
            # Stop at first row starting with "None" value
            if not row[0].value:
                break

            new_dict = {}
            for i in range(nb_keys):
                new_dict[key_list[i]] = row[i].value

            data_list.append(new_dict)

        return data_list

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def check_next_row_element(self, row, col):
        row_items = []
        for val in self.values:
            if val.row == row and val.col >= col:
                row_items.append(val)

        return row_items

    def check_row_header(self, col):
        for val in self.row_values:
            if val.row == self.first_row_idx and val.col == col:
                return val.value

    def get_next_row(self, val):
        value_len = len(self.values)
        for i in range(value_len):
            if val == self.values[i].value and i + 1 <= value_len:
                row_item = {}
                next_row_item = self.values[i + 1]
                col = next_row_item.col
                row = next_row_item.row
                next_items = self.check_next_row_element(row, col)

                for next_item in next_items:
                    header = self.check_row_header(next_item.col)
                    row_item[header] = next_item.value

        return row_item

    def reset(self):
        self.counter = 0
        self.row_counter = 0


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            next = excel_reader.get_next()

    return None


def update_dictionary(dic):
    for key in dic:
        nest_dict = dic[key]
        for nest_key in nest_dict.keys():
            next_item = excel_reader.get_next_row(nest_key)
            nest_dict[nest_key] = next_item

        # print(dict.keys())
    return dic


d = {}

excel_reader = ExcelReader()
root = excel_reader.get_next()

generate_dictionary(d, root)

dic = update_dictionary(d)

print(" dic: {}".format(dic))

# with open('types_output.json', 'w') as json_file:
#     json.dump(dic, json_file, indent=4)
