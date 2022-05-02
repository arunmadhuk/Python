import openpyxl
import json


# Working class to contain an "item" (value, column)
class Item:
    def __init__(self, value, col):
        self.value = value
        self.col = col


# Excel reader class
class ExcelReader:
    # TODO: don't set excel file name a default
    # TODO: add main sheet name as param?
    def __init__(self, excel_file="template.xlsx"):
        self.counter = 0
        self.values = []

        self.work_book = openpyxl.load_workbook(excel_file)
        self.get_values()

    def get_values(self, sheet_name="MAIN"):
        if sheet_name not in self.work_book.sheetnames:
            return []

        sheet = self.work_book[sheet_name]
        data_range = sheet.calculate_dimension()
        data = sheet[data_range]

        # Check that first cell is not empty
        # TODO: should check if is "root" or "project" or something specific...
        if not data[0][0].value:
            # TODO: print better error
            print("Error: first cell not root")
            return

        first_column_idx = data[0][0].column

        self.values = []
        for row in data:
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    col = cell.column - first_column_idx
                    self.values.append(Item(value, col))
                    # Only 1 value per row, rest is ignored
                    break

        return self.values

    def load_sheet(self, sheet_name):
        if sheet_name not in self.work_book.sheetnames:
            return []

        selected_sheet = self.work_book[sheet_name]
        print(selected_sheet)
        data_range = selected_sheet.calculate_dimension()
        print(data_range)
        data = selected_sheet[data_range]

        # Get column titles
        key_list = []
        first = True
        for column_title in data[0]:
            print(column_title.value)
            # Stop at first "None" value
            if not column_title.value:
                break

            # Special case for first column (ID)
            if not first:
                key_list.append(column_title.column_letter)
            else:
                key_list.append("[ID]")
                first = False
        print(key_list)
        nb_keys = len(key_list)

        # Create dictionary for each row of table
        data_list = []
        for row in data:
            # Stop at first row starting with "None" value
            if not row[0].value:
                break

            new_dict = {}
            for i in range(nb_keys):
                new_dict[key_list[i]] = row[i].value

            data_list.append(new_dict)

        return data_list

    def get_next(self):
        item = None
        if self.counter < len(self.values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def reset(self):
        self.counter = 0

    def get_next_type(self):
        item = None
        if self.counter < len(self.types_values):
            item = self.values[self.counter]
            self.counter += 1

        return item

    def load_types_sheet(self):

        sheet = self.work_book["TYPES"]
        data_range = sheet.calculate_dimension()
        print(data_range)
        data = sheet[data_range]

        # Check that first cell is not empty
        # TODO: should check if is "root" or "project" or something specific...
        if not data[0][0].value:
            # TODO: print better error
            print("Error: first cell not root")

        first_column_idx = data[0][0].column
        first_row_idx = data[0][0].row
        print(first_column_idx)
        self.types_values = []
        for row in data:
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    col = cell.column - first_column_idx
                    # print("value  {} col {}".format(value, col))
                    row = cell.row - first_row_idx
                    print("value -- {} -- row {} -- col {}".format(value, row, col))
                    self.types_values.append(Item(value, col))
                    # Only 1 value per row, rest is ignored
                    # break

        print(self.types_values)


# Process to generate the dictionary
def generate_dictionary(dic, prev):
    next = excel_reader.get_next()

    while next:
        if next.col < prev.col:
            return next

        if next.col == prev.col:
            dic[next.value] = {}
            prev = next
            next = excel_reader.get_next()

        elif next.col == prev.col + 1:
            dic[prev.value] = {next.value: {}}
            next = generate_dictionary(dic[prev.value], next)
        else:
            next = excel_reader.get_next()

    return None


d = {}

excel_reader = ExcelReader()
root = excel_reader.get_next()

excel_reader.load_types_sheet()
generate_dictionary(d, root)
print(" d: {}".format(d))


# print("Types {}".format(types_list))
